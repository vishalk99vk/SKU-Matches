"""
Reads the uploaded workbook (Client_Data + AIAS sheets), matches every AIAS
row against every Client_Data row by image similarity, and writes an output
workbook with three sheets:

  1. Matched          - SKU_Name | Confidence_Score | AIAS_Group_Name | AIAS_Image_Link
  2. Unmatched_AIAS    - AIAS rows with no good match in Client_Data ("clustered")
  3. Unmatched_Client  - Client SKUs with no good match in AIAS ("not trained")

CONFIDENCE_THRESHOLD decides what counts as a real match vs. "no match found".
Tune this once you see real scores on your data.
"""

import openpyxl
from openpyxl.styles import Font, PatternFill
from image_matcher import compare_images

CONFIDENCE_THRESHOLD = 55.0  # 0-100, tune based on real results


def _read_sheet_pairs(ws):
    """Yields (name, image_source) for every populated data row after the header."""
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row and row[0] and row[1]:
            rows.append((str(row[0]).strip(), str(row[1]).strip()))
    return rows


def run_matching(input_path: str, output_path: str, progress_callback=None):
    wb = openpyxl.load_workbook(input_path)

    if "Client_Data" not in wb.sheetnames or "AIAS" not in wb.sheetnames:
        raise ValueError("Workbook must contain 'Client_Data' and 'AIAS' sheets.")

    client_rows = _read_sheet_pairs(wb["Client_Data"])
    aias_rows = _read_sheet_pairs(wb["AIAS"])

    matched = []          # (sku_name, confidence, aias_name, aias_image)
    matched_client_names = set()
    matched_aias_names = set()

    total = len(aias_rows)
    for i, (aias_name, aias_image) in enumerate(aias_rows):
        best_score = -1.0
        best_client = None

        for client_name, client_image in client_rows:
            try:
                score = compare_images(client_image, aias_image)
            except Exception as e:
                print(f"  [warn] could not compare {client_name} vs {aias_name}: {e}")
                continue

            if score > best_score:
                best_score = score
                best_client = (client_name, client_image)

        if best_client and best_score >= CONFIDENCE_THRESHOLD:
            matched.append((best_client[0], best_score, aias_name, aias_image))
            matched_client_names.add(best_client[0])
            matched_aias_names.add(aias_name)

        if progress_callback:
            progress_callback(i + 1, total)

    unmatched_aias = [(n, img) for n, img in aias_rows if n not in matched_aias_names]
    unmatched_client = [(n, img) for n, img in client_rows if n not in matched_client_names]

    _write_output(output_path, matched, unmatched_aias, unmatched_client)
    return {
        "matched": len(matched),
        "unmatched_aias": len(unmatched_aias),
        "unmatched_client": len(unmatched_client),
    }


def _style_header(ws, headers):
    fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    font = Font(bold=True, color="FFFFFF")
    for col, h in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = fill
        cell.font = font
    for col_idx in range(1, len(headers) + 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = 35


def _write_output(output_path, matched, unmatched_aias, unmatched_client):
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    ws1 = wb.create_sheet("Matched")
    _style_header(ws1, ["SKU_Name", "Confidence_Score", "AIAS_Group_Name", "AIAS_Image_Link"])
    for r, (sku, score, aias_name, aias_img) in enumerate(matched, start=2):
        ws1.cell(row=r, column=1, value=sku)
        ws1.cell(row=r, column=2, value=score)
        ws1.cell(row=r, column=3, value=aias_name)
        ws1.cell(row=r, column=4, value=aias_img)

    ws2 = wb.create_sheet("Unmatched_AIAS")
    _style_header(ws2, ["AIAS_Group_Name", "AIAS_Image_Link", "Status"])
    for r, (name, img) in enumerate(unmatched_aias, start=2):
        ws2.cell(row=r, column=1, value=name)
        ws2.cell(row=r, column=2, value=img)
        ws2.cell(row=r, column=3, value="Clustered - not present in Client_Data")

    ws3 = wb.create_sheet("Unmatched_Client")
    _style_header(ws3, ["SKU_Name", "Image_Link", "Status"])
    for r, (name, img) in enumerate(unmatched_client, start=2):
        ws3.cell(row=r, column=1, value=name)
        ws3.cell(row=r, column=2, value=img)
        ws3.cell(row=r, column=3, value="NA - not trained in this round of AI training")

    wb.save(output_path)
