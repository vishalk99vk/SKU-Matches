"""
Generates the template Excel file that users download from the web app
and fill in with their own data before uploading.
"""
import openpyxl
from openpyxl.styles import Font, PatternFill

wb = openpyxl.Workbook()
wb.remove(wb.active)

header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
header_font = Font(bold=True, color="FFFFFF")

# --- Sheet 1: Client_Data (the trusted reference catalog) ---
client = wb.create_sheet("Client_Data")
client["A1"] = "SKU_Name"
client["B1"] = "Image_Link_or_Path"
client["A2"] = "Mango_12mg"
client["B2"] = "https://example.com/images/mango_12mg.jpg"
client["A3"] = "Mango_15mg"
client["B3"] = "https://example.com/images/mango_15mg.jpg"
client["A4"] = "Berry_20mg"
client["B4"] = "https://example.com/images/berry_20mg.jpg"
for c in ["A1", "B1"]:
    client[c].fill = header_fill
    client[c].font = header_font
client.column_dimensions["A"].width = 22
client.column_dimensions["B"].width = 45

# --- Sheet 2: AIAS (the low quality / randomly named data to match) ---
aias = wb.create_sheet("AIAS")
aias["A1"] = "Group_Name"
aias["B1"] = "Image_Link_or_Path"
aias["A2"] = "Group_001"
aias["B2"] = "https://example.com/images/group_001.jpg"
aias["A3"] = "Group_002"
aias["B3"] = "https://example.com/images/group_002.jpg"
aias["A4"] = "Group_003"
aias["B4"] = "https://example.com/images/group_003.jpg"
for c in ["A1", "B1"]:
    aias[c].fill = header_fill
    aias[c].font = header_font
aias.column_dimensions["A"].width = 22
aias.column_dimensions["B"].width = 45

wb.save("static/Image_Matching_Template.xlsx")
print("Template saved to static/Image_Matching_Template.xlsx")
